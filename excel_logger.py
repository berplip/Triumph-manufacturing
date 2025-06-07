# excel_logger.py
# Funcionalidad para registrar la actividad de la aplicación en un archivo Excel.

import os
import datetime
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
import threading

# Importa desde config y data_manager
from config import EXCEL_LOG_FILE_FULL_PATH, LOGS_PATH
import data_manager 

# MODIFICADO: Se añade un Lock para evitar escrituras simultáneas que puedan corromper el archivo.
excel_lock = threading.Lock()

def inicializar_excel_log():
    """Crea el archivo Excel y la carpeta de logs si no existen."""
    # Esta función no necesita cambios, pero se mantiene por completitud.
    if not os.path.exists(LOGS_PATH):
        try:
            os.makedirs(LOGS_PATH, exist_ok=True)
        except OSError as e:
            print(f"ADVERTENCIA: No se pudo crear el directorio de logs '{LOGS_PATH}': {e}")

    if not os.path.exists(EXCEL_LOG_FILE_FULL_PATH):
        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Registro de Actividad"
            headers = ["Timestamp", "Usuario", "Rol", "Accion", "Detalles Adicionales", "Duracion Sesion (min)"]
            sheet.append(headers)
            for col_num, header_title in enumerate(headers, 1):
                cell = sheet.cell(row=1, column=col_num)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")
                column_letter = get_column_letter(col_num)
                if header_title == "Timestamp": sheet.column_dimensions[column_letter].width = 20
                elif header_title == "Detalles Adicionales": sheet.column_dimensions[column_letter].width = 45
                elif header_title == "Duracion Sesion (min)": sheet.column_dimensions[column_letter].width = 22
                else: sheet.column_dimensions[column_letter].width = 25
            workbook.save(EXCEL_LOG_FILE_FULL_PATH)
            workbook.close()
        except Exception as e:
            print(f"ADVERTENCIA: No se pudo inicializar el archivo de log Excel '{EXCEL_LOG_FILE_FULL_PATH}': {e}")

def registrar_accion_excel(accion, detalles="", duracion_min=None):
    """Registra una acción en el archivo Excel de forma segura."""
    # Se usa un Lock para asegurar que solo un hilo escriba a la vez.
    with excel_lock:
        try:
            # Cargar o crear el libro de trabajo
            if os.path.exists(EXCEL_LOG_FILE_FULL_PATH):
                workbook = openpyxl.load_workbook(EXCEL_LOG_FILE_FULL_PATH)
                sheet = workbook.active
            else: 
                # Si el archivo no existe por alguna razón, se reinicializa.
                inicializar_excel_log()
                workbook = openpyxl.load_workbook(EXCEL_LOG_FILE_FULL_PATH)
                sheet = workbook.active
        
            timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            nombre_usuario_log = data_manager.usuario_actual.get("nombre", "N/A")
            rol_usuario_log = data_manager.usuario_actual.get("rol", "N/A")
            
            # Asegura que la entrada siempre tenga el número correcto de columnas.
            duracion_str = f"{duracion_min:.2f}" if duracion_min is not None else ""
            log_entry = [timestamp, nombre_usuario_log, rol_usuario_log, accion, detalles, duracion_str]
            
            sheet.append(log_entry)
            workbook.save(EXCEL_LOG_FILE_FULL_PATH)
            workbook.close()

        except Exception as e:
            print(f"ERROR al registrar acción en Excel: {e}. Acción: '{accion}', Detalles: '{detalles}'")